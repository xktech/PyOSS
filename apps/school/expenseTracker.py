import os
import json
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import csv

def expenseTracker(): # Added from Student Dashboard
    EXPENSE_FILE = "expenses.json"
    def loadExpenses():
        if os.path.exists(EXPENSE_FILE):
            with open(EXPENSE_FILE, "r") as f:
                return json.load(f)
        return []
    
    expenses = loadExpenses()

        
    def saveExpenses():
        with open(EXPENSE_FILE, "w") as f:
            json.dump(expenses, f, indent=2)

    def get_number(prompt):
        while True:
            try:
                return int(input(prompt))
            except ValueError:
                print("Please enter a valid number!")

    def addExpense():
        usrExpense = input("Enter expense: ")
        expenseCost = float(input("Enter the cost of the item: £"))
        expenseCatergory = input("Enter the Category (Food, Work etc): ")
        date_added = datetime.now().strftime("%d/%m/%Y")

        expenses.append({
            "Expense": usrExpense,
            "Cost": expenseCost,
            "Catergory": expenseCatergory,
            "Date": date_added,
        })

        saveExpenses()

    def showMonthlyExpenses():
        now = datetime.now()
        total = 0

        print("\nThis Month's Expenses:\n")

        for expense in expenses:
            expense_date = datetime.strptime(expense["Date"], "%d/%m/%Y")

            if expense_date.month == now.month and expense_date.year == now.year:
                print(f"{expense['Expense']} - £{expense['Cost']} | {expense['Date']}")
                total += expense["Cost"]

        print(f"\nTotal this month: £{total:.2f}\n")

    def exportExpense():
        sorted_expenses = sorted(expenses, key=lambda x: x["Cost"])

        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"

        # Header row
        headers = ["Expense", "Cost (£)", "Catergory", "Date"]
        ws.append(headers)

        # Header style
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_font = Font(color="FFFFFF", bold=True)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        row_num = 2
        for expense in sorted_expenses:
            ws.append([
                expense["Expense"],
                expense["Cost"],
                expense["Catergory"],
                expense["Date"]
            ])

            # Alternate row colours
            fill = PatternFill("solid", fgColor="DCE6F1") if row_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

            for col in range(1, 5):
                ws.cell(row=row_num, column=col).fill = fill
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="center")

            # Currency format
            ws.cell(row=row_num, column=2).number_format = '£#,##0.00'

            row_num += 1

        # Auto column widths
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column

            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    pass

            adjusted_width = max_length + 4
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width

        # Freeze top row
        ws.freeze_panes = "A2"

        wb.save("expenses.xlsx")
        print("Spreadsheet exported!")

    while True:
        print("1. Add expense")
        print("2. Export to Excel")
        print("3. Show monthly expenses")
        print("4. Quit")

        choice = get_number("Choose: ")

        if choice == 1:
            addExpense()

        elif choice == 2:
            exportExpense()

        elif choice == 3:
            showMonthlyExpenses()

        elif choice == 4:
            break