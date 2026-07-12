"""
    This will have: 
    1. A login / logout system
    2. A notes system
    3. Timer
    4. Music (Hopefully)
    5. A calculator
    6. File system
    7. More in the future"""

import os
import sys
import time
import json
import random
import datetime
import getpass
import subprocess
import platform
from pathlib import Path
from colorama import init, Fore, Style
import winsound
import psutil
import threading
import spotipy
from spotipy.oauth2 import SpotifyClientCredentials
from spotipy.oauth2 import SpotifyOAuth
from dotenv import load_dotenv
import shutil
import curses
from datetime import datetime
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# For GUI Later
import tkinter as tk
from tkinter import messagebox

# TOP FUNCTIONS
def get_number(prompt):
    while True:
        try:
            return int(input(prompt))
        except:
            print("Please enter a valid number!")
load_dotenv()
init(autoreset=True)

def play_async(sound):
    threading.Thread(target=lambda: winsound.PlaySound(sound, winsound.SND_ALIAS), daemon=True).start()
def center_text(text=""):
    width = shutil.get_terminal_size().columns
    return text.center(width)

def print_centered(text=""):
    print(center_text(text))

def center_block(text):
    """Centers a multi-line block (like ASCII art) as one unit,
    so internal alignment isn't destroyed."""
    width = shutil.get_terminal_size().columns
    lines = text.split("\n")
    max_len = max(len(line) for line in lines)
    pad = max(0, (width - max_len) // 2)
    return "\n".join((" " * pad) + line for line in lines)

def center_input_prompt(text=""):
    """Like center_text, but for use with input() —
    only pads on the left, so the cursor lands in the middle,
    not pushed off to the right by trailing spaces."""
    width = shutil.get_terminal_size().columns
    pad = max(0, (width - len(text)) // 2)
    return " " * pad + text



# SOUNDS
def error_sound():
    winsound.PlaySound("SystemHand", winsound.SND_ALIAS)
def success_sound():
    winsound.Beep(800, 100)

# LOGIN SYSTEM
USERS_FILE = "users.json"
def load_users():
    if not os.path.exists(USERS_FILE):
        return {}
    try:
        with open(USERS_FILE, "r") as f:
            return json.load(f)
    except json.JSONDecodeError:
        return {}
def save_users(users):
    with open(USERS_FILE, "w") as f:
        json.dump(users, f, indent=4)
def register_user():
    print("\n--- REGISTER NEW USER ---")

    users = load_users()

    while True:
        username = input("Choose username: ").strip()

        if username in users:
            print("Username already exists. Try another.")
            continue
        if username == "":
            print("Username cannot be empty / have spaces in.")
            continue
        break

    while True:
        password = input("Create password: ").strip()

        if len(password) < 3:
            print("Password too short (min 3 chars).")
            continue
        break

    users[username] = password
    save_users(users)

    print(f"\nUser '{username}' created successfully!")
    return username
def login_user():
    print("\n--- LOGIN ---")

    users = load_users()

    username = input("Username: ").strip()
    password = getpass.getpass("Password: ")

    if username in users and users[username] == password:
        print(f"\nWelcome back, {username}!")
        return username

    print("Invalid username or password.")
    return None
def shutdown():
    new_window(username)
    print_centered(Fore.RED + "Shutting down..")
    print_centered()
    time.sleep(1)
    print_centered(Fore.CYAN + f"Goodbye {username}!")
    print_centered()
    time.sleep(0.5)
    sys.exit(0)

# System Stuff
def systemSpecs():
    new_window(username)
    while True:
        print("System:", platform.system())
        print("Release:", platform.release())
        print("Version:", platform.version())
        print("Machine:", platform.machine())
        print("Processor:", platform.processor())

        print("CPU Cores:", psutil.cpu_count(logical=True))
        print("RAM GB:", round(psutil.virtual_memory().total / (1024**3), 2))
        
        break
def files():
    new_window(username)
    
    USERNAME = username


    os.makedirs("users", exist_ok=True)

    SAVE_FILE = f"users/{USERNAME}_files.json"

    default_fs = {
        "home": {
            "notes.txt": "Hello World!",
            "games": {},
            "system": {
                "config.sys": "darkmode=true"
            }
        }
    }

    # LOAD FILESYSTEM
    if os.path.exists(SAVE_FILE):
        with open(SAVE_FILE, "r") as f:
            file_system = json.load(f)
    else:
        file_system = default_fs

        with open(SAVE_FILE, "w") as f:
            json.dump(file_system, f, indent=4)

    # SAVE FUNCTION
    def save_fs():
        with open(SAVE_FILE, "w") as f:
            json.dump(file_system, f, indent=4)

    # LOAD FILESYSTEM
    if os.path.exists(SAVE_FILE):
        with open(SAVE_FILE, "r") as f:
            file_system = json.load(f)
    else:
        file_system = default_fs

        with open(SAVE_FILE, "w") as f:
            json.dump(file_system, f, indent=4)

    path = ["home"]

    def get_current_dir():
        current = file_system
        for folder in path:
            current = current[folder]
        return current

    def pwd():
        return "/" + "/".join(path)

    print("You are in the home directory.")
    print("Commands: ls, cd <folder>, cd .., pwd, exit, cat, mkdir, touch, rm")

    while True:
        cmd = input("> ").strip()
        parts = cmd.split()

        if len(parts) == 0:
            continue

        command = parts[0]

        # EXIT
        if command == "exit":
            print("Exiting...")
            time.sleep(1)
            break

        # LS
        elif command == "ls":
            current = get_current_dir()
            for item in current:
                print(item)

        # PWD
        elif command == "pwd":
            print(pwd())

        # CD
        elif command == "cd":
            if len(parts) < 2:
                print("Usage: cd <folder> or cd ..")
                continue

            target = parts[1]
            current = get_current_dir()

            if target == "..":
                if len(path) > 1:
                    path.pop()
            elif target in current and isinstance(current[target], dict):
                path.append(target)
            else:
                print("Folder not found")
        
        # CAT
        elif command == "cat":
            if len(parts) < 2:
                print("Usage: cat <file>")
                continue

            filename = parts[1]
            current = get_current_dir()

            if filename in current and isinstance(current[filename], str):
                print(current[filename])
            else:
                print("File not found.")


        # MKDIR
        elif command == "mkdir":
            if len(parts) < 2:
                print("Usage: mkdir <path>")
                return

            parts_path = parts[1].split("/")
            folder = get_current_dir()

            for p in parts_path[:-1]:
                if p not in folder or not isinstance(folder[p], dict):
                    print(f"mkdir: cannot create '{parts[1]}': No such directory")
                    return
                folder = folder[p]

            new_dir = parts_path[-1]

            if new_dir in folder:
                print("mkdir: File exists")
                return

            folder[new_dir] = {}
            save_fs()
            print(f"Created folder: {parts[1]}")

        # TOUCH
        elif command == "touch":
            if len(parts) < 2:
                print("Usage: touch <filename>")
                return
            
            filename = parts[1]
            folder = get_current_dir()

            if filename in folder:
                print(f"touch: '{filename}' already exists")
                return
            
            folder[filename] = ""
            save_fs()
            print(f"Touched: {filename}")


        # RM
        elif command == "rm":
            if len(parts) < 2:
                print("Usage: rm <file/folder>")
                return
            
            name = parts[1]
            folder = get_current_dir()

            if name not in folder:
                print(f"rm: cannot remove '{name}': No such file or directory.")
                return
            
            del folder[name]
            save_fs()
            print(f"Removed: {name}")

        # EDIT FILES
        elif command == "edit":
            if len(parts) < 2:
                print("Usage: edit <filename>")
                return
            
            filename = parts[1]
            folder = get_current_dir()

            if filename not in folder:
                print(f"edit: '{filename}' does not exist.")
                return
            
            print("\n--- EDIT MODE ---")
            print("Type your content below.")
            print("Type ':sq' to save and quit.")
            print("NOTE: It erases the current contents of the file. You will have to")
            print("re-write it again. This is a bug but not an important one currently\n")

            print(f"Current content: \n{folder[filename]}\n")


            lines = []

            while True:
                line = input()

                if line == ":sq":
                    break

                lines.append(line)

            folder[filename] = "\n".join(lines)
            save_fs()
            print(f"Saved: {filename}")

# Other
def calculator():
    new_window(username)
    while True:
        OPERATORS = ["+", "-", "*", "/", "exit"]
        usrOp = input(f"Enter the operator from {OPERATORS} ")

        if usrOp not in OPERATORS:
            print(f"Please chose from {OPERATORS} ")
            continue
        
        elif usrOp == "exit":
            print("Exiting")
            time.sleep(1)
            break

        else:
            num1 = int(input("Enter x (number 1): "))
            num2 = int(input("Enter y (number 2): "))

            if usrOp == "*":
                answer = num1 * num2
                print(f"{num1} * {num2} = {answer}")
            elif usrOp == "/":
                answer = num1 / num2
                print(f"{num1} / {num2} = {answer}")
            elif usrOp == "+":
                answer = num1 + num2
                print(f"{num1} + {num2} = {answer}")
            elif usrOp == "-":
                answer = num1 - num2
                print(f"{num1} - {num2} = {answer}")
def notes():
    new_window(username)
    NOTES_FILE = "notes.json"
    
    def loadNotes():
        if os.path.exists(NOTES_FILE):
            with open(NOTES_FILE, "r") as f:
                return json.load(f)
            return []
        
    notes = loadNotes()

    def saveNotes():
        with open(NOTES_FILE, "w") as f:
            json.dump(notes, f, indent=2)

    def viewNotes():
        if not notes:
            print("No notes to display!")
            return

        print("\n--- NOTES ---")
        for i, nts in enumerate(notes, 1):
            print(f"{i}. Note: {nts['Note']} | {nts['Date']}")
        print()

    def addNote():
        note = input("Enter the note: ")
        
        notes.append({
            "Note": note,
            "Date": datetime.now().strftime("%d/%m/%Y"),
        })
        saveNotes()

    while True:
        print("\n1. Add Note")
        print("2. View Notes")
        print("3. Exit")

        choice = get_number("Choose: ")

        if choice == 1:
            addNote()   
        elif choice == 2:
            viewNotes()
        elif choice == 3:
            break
def music():
    new_window(username)
    CLIENT_ID = os.getenv("CLIENT_ID")
    CLIENT_SECRET = os.getenv("CLIENT_SECRET")

    sp = spotipy.Spotify(auth_manager=SpotifyOAuth(
        client_id=CLIENT_ID,
        client_secret=CLIENT_SECRET,
        redirect_uri="http://127.0.0.1:8888/callback",
        scope="user-read-playback-state,user-modify-playback-state"
    ))

    def spotifySearch():
        song = input("Song name: ")

        results = sp.search(q=song, limit=5, type='track')

        print("\nResults:\n")

        for i, track in enumerate(results['tracks']['items'], start=1):
            name = track['name']
            artist = track['artists'][0]['name']

            print(f"{i}. {name} - {artist}")

    def playSong(song_name):

        results = sp.search(q=song_name, limit=1, type='track')

        tracks = results['tracks']['items']

        if not tracks:
            print("Song not found.")
            return

        track = tracks[0]
        uri = track['uri']

        devices = sp.devices()
        for d in devices['devices']:   
            print(Fore.GREEN + "--- DEBUG ---") 
            print(Fore.GREEN + d['name'], '| active:', d['is_active'], '| restricted:', d['is_restricted'])
            print(Fore.GREEN + "-------------")
        if not devices['devices']:
            print("No active Spotify device found. Please open spotify.")
            return

        device_id = devices['devices'][0]['id']

        sp.start_playback(device_id=device_id, uris=[uri])

        print(f"Now playing: {track['name']}")

    def get_active_device_id():
        devices = sp.devices()

        if not devices['devices']:
            return None

        for d in devices['devices']:
            if d['is_active']:
                return d['id']

        # nothing marked active — fall back to first device
        return devices['devices'][0]['id']

    while True:
        print("\n=== Spotify ===")
        print("1. Search for a song")
        print("2. Play a song")
        print("3. Pause")
        print("4. Resume")
        print("5. Skip")
        print("6. Exit")

        choice = input(">>> ")

        if choice == "1":
            spotifySearch()

        elif choice == "2":
            song_choice = input("Enter a song name: ")
            playSong(song_choice)

        elif choice == "3":
            device_id = get_active_device_id()
            if device_id is None:
                print("No active Spotify device found. Please open Spotify.")
            else:
                sp.pause_playback(device_id=device_id)
                print("Paused.")

        elif choice == "4":
            device_id = get_active_device_id()
            if device_id is None:
                print("No active Spotify device found. Please open Spotify.")
            else:
                sp.start_playback(device_id=device_id)
                print("Resumed.")

        elif choice == "5":
            sp.next_track()
            print("Skipped.")

        elif choice == "6":
            break

# Games
def random_number_game():
    number = random.randint(1, 100)
    print("In this game. You guess a number (1 - 100 ('123' to leave))")
    while True:
        guess = get_number("Enter Your Guess: ")

        if guess == 123:
            break
        elif guess == number:
            print("You got it correct!")
            keep_going = input("Keep Going? (Y/N) ").lower()
            if keep_going == "y":
                number = random.randint(1, 100)  # reset the number
                continue
            elif keep_going == "n":
                break
            else:
                print("Only Y / N")
        else:
            print("You got it wrong!")
def rps():
    # Rock Paper Scissors
    OPTIONS = ['rock', 'paper', 'scissors']

    beats = {
            "paper":"rock",
            "scissors":"paper",
            "rock":"scissors",
        }


    while True:
        ai_choice = random.choice(OPTIONS)
        
        human_choice = input("Rock, Paper or Scissors? (q to exit): ").lower()


        print(f"Your Choice: {human_choice}")
        print(f"AI Choice: {ai_choice}")

        if human_choice == "q":
            break
        else:
            if human_choice == ai_choice:
                print("TIE!")
            elif beats[human_choice] == ai_choice:
                print("You WIN!")
            else:
                print("AI Wins!")
def higher_lower_game(low=1, high=100, max_attempts=10):
    target = random.randint(low, high)
    attempts = 0

    print(f"Guess a number between {low} and {high}!")

    while attempts < max_attempts:
        attempts += 1
        guess = int(input(f"Attempt {attempts}/{max_attempts}: "))

        if guess == target:
            print(f"Correct! The number was {target}.")
            return True
        elif guess < target:
            print("Higher!")
        else:
            print("Lower!")

    print(f"Out of attempts! The number was {target}.")
    return False
def pong():
    # Completely Vibe Coded. Couldnt be asked to code this
    WIDTH = 60
    HEIGHT = 20
    PADDLE_SIZE = 4

    def game(stdscr):
        curses.curs_set(0)
        stdscr.nodelay(True)
        stdscr.timeout(20)

        left_y = HEIGHT // 2 - PADDLE_SIZE // 2
        right_y = HEIGHT // 2 - PADDLE_SIZE // 2

        ball_x = WIDTH // 2
        ball_y = HEIGHT // 2
        dx = 1
        dy = 1

        left_score = 0
        right_score = 0

        while True:
            key = stdscr.getch()

            # Controls
            if key == ord("w") and left_y > 2:
                left_y -= 1
            elif key == ord("s") and left_y < HEIGHT - PADDLE_SIZE:
                left_y += 1
            elif key == ord("q"):
                break
            # Simple AI
            ai_center = right_y + PADDLE_SIZE // 2

            if ball_y < ai_center and right_y > 2:
                right_y -= 1
            elif ball_y > ai_center and right_y < HEIGHT - PADDLE_SIZE:
                right_y += 1

            # Move ball
            ball_x += dx
            ball_y += dy

            # Bounce off top/bottom
            if ball_y <= 2 or ball_y >= HEIGHT - 1:
                dy *= -1

            # Left paddle collision / score
            if ball_x == 2:
                if left_y <= ball_y < left_y + PADDLE_SIZE:
                    dx *= -1
                else:
                    right_score += 1
                    ball_x = WIDTH // 2
                    ball_y = HEIGHT // 2
                    dx = 1

            # Right paddle collision / score
            if ball_x == WIDTH - 3:
                if right_y <= ball_y < right_y + PADDLE_SIZE:
                    dx *= -1
                else:
                    left_score += 1
                    ball_x = WIDTH // 2
                    ball_y = HEIGHT // 2
                    dx = -1

            # Draw
            stdscr.clear()
            stdscr.addstr(
                0, 0,
                f"Left: {left_score}  Right: {right_score}   W/S to go up and down  Q=Quit"
            )

            # Top & bottom borders
            for x in range(WIDTH):
                stdscr.addch(1, x, "-")
                stdscr.addch(HEIGHT, x, "-")

            # Paddles
            for i in range(PADDLE_SIZE):
                stdscr.addch(left_y + i, 1, "|")
                stdscr.addch(right_y + i, WIDTH - 2, "|")

            # Ball
            stdscr.addch(ball_y, ball_x, "O")

            stdscr.refresh()
            time.sleep(0.02)

            

    curses.wrapper(game)
def text_adventure():
    #TODO: do this shit
    pass


# SCHOOL
def expenseTracker(): # Added from Student Dashboard
    EXPENSE_FILE = "expenses.json"
    def loadExpenses():
        if os.path.exists(EXPENSE_FILE):
            with open(EXPENSE_FILE, "r") as f:
                return json.load(f)
        return []
    
    expenses = loadExpenses()

        
    def saveExpenses():
        with open(EXPENSE_FILE, "w") as f:
            json.dump(expenses, f, indent=2)

    def get_number(prompt):
        while True:
            try:
                return int(input(prompt))
            except ValueError:
                print("Please enter a valid number!")

    def addExpense():
        usrExpense = input("Enter expense: ")
        expenseCost = float(input("Enter the cost of the item: £"))
        expenseCatergory = input("Enter the Category (Food, Work etc): ")
        date_added = datetime.now().strftime("%d/%m/%Y")

        expenses.append({
            "Expense": usrExpense,
            "Cost": expenseCost,
            "Catergory": expenseCatergory,
            "Date": date_added,
        })

        saveExpenses()

    def showMonthlyExpenses():
        now = datetime.now()
        total = 0

        print("\nThis Month's Expenses:\n")

        for expense in expenses:
            expense_date = datetime.strptime(expense["Date"], "%d/%m/%Y")

            if expense_date.month == now.month and expense_date.year == now.year:
                print(f"{expense['Expense']} - £{expense['Cost']} | {expense['Date']}")
                total += expense["Cost"]

        print(f"\nTotal this month: £{total:.2f}\n")

    def exportExpense():
        sorted_expenses = sorted(expenses, key=lambda x: x["Cost"])

        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"

        # Header row
        headers = ["Expense", "Cost (£)", "Catergory", "Date"]
        ws.append(headers)

        # Header style
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_font = Font(color="FFFFFF", bold=True)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        row_num = 2
        for expense in sorted_expenses:
            ws.append([
                expense["Expense"],
                expense["Cost"],
                expense["Catergory"],
                expense["Date"]
            ])

            # Alternate row colours
            fill = PatternFill("solid", fgColor="DCE6F1") if row_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

            for col in range(1, 5):
                ws.cell(row=row_num, column=col).fill = fill
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal="center")

            # Currency format
            ws.cell(row=row_num, column=2).number_format = '£#,##0.00'

            row_num += 1

        # Auto column widths
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column

            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = max_length + 4
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width

        # Freeze top row
        ws.freeze_panes = "A2"

        wb.save("expenses.xlsx")
        print("Spreadsheet exported!")

    while True:
        print("1. Add expense")
        print("2. Export to Excel")
        print("3. Show monthly expenses")
        print("4. Quit")

        choice = get_number("Choose: ")

        if choice == 1:
            addExpense()

        elif choice == 2:
            exportExpense()

        elif choice == 3:
            showMonthlyExpenses()

        elif choice == 4:
            break
def homeworkList(): # Added from Student Dashboard





    HOMEWORK_FILE = "homework.json"

    def loadHomework():
        if os.path.exists(HOMEWORK_FILE):
            with open(HOMEWORK_FILE, "r") as f:
                return json.load(f)
        return []
    homework = loadHomework() 

    def getValidDate(prompt):
        while True:
            date_str = input(prompt)
            try:
                datetime.strptime(date_str, "%d/%m/%Y")
                return date_str
            except ValueError:
                print("Invalid date! Please use: DD/MM/YYYY format (e.g. 28/04/2026)")

    def saveHomework():
        with open(HOMEWORK_FILE, "w") as f:
            json.dump(homework, f, indent=2)
    def addHomework():
        subject = input("Enter the subject: ")
        teacher = input("Enter the teacher: ")
        task = input("Enter your homework task: ")
        due_date = getValidDate("Enter due date (DD/MM/YYYY): ")

        homework.append({
            "Subject": subject,
            "Teacher": teacher,
            "Task": task,
            "Due": due_date,
            "Date": datetime.now().strftime("%d/%m/%Y")
        })
        saveHomework()
    def viewHomework():
        def viewHomework():
            if not homework:
                print("No homework saved!")
                return

        print("\n--- Homework List ---")
        for i, hw in enumerate(homework, 1):
            print(f"{i}. [{hw['Subject']}] {hw['Task']} - {hw['Teacher']} | Due: {hw['Due']} {hw['Date']}")
        print()

    while True:
        print("\n1. Add Homework:")
        print("2. View Homework")
        print("3. Exit")

        choice = get_number("Choose: ")

        if choice == 1:
            addHomework()
        if choice == 2:
            viewHomework()
        if choice == 3:
            break


# HOMEPAGE
def render_home(username):  
    print("-" * 38)
    print("Files                      /GAMES")
    print("System Specs               /SCHOOL")
    print("Calculator                 Placeholder")
    print("Music                      Placeholder")
    print("Notes                      Placeholder")
    print("         Logout or Shutdown           ")
    print("-" * 38)


def render_games(username):
    print("-" * 38)
    print("1. Number Game")
    print("2. Rock Paper Scissors")
    print("3. Higher or Lower")
    print("4. Pong")
    print("5. Type 'back' to return")
    print("-" * 38)


def render_school(username):
    print("-" * 38)
    print("1. Expenses Tracker")
    print("2. Homework List")
    print("")

def games_menu(username):
    while True:
        render_games(username)

        gameChoice = input(">> ").lower().strip()

        if gameChoice == "number game":
            random_number_game()
        elif gameChoice == "rock paper scissors":
            rps()
        elif gameChoice == "pong":
            if __name__ == "__main__":
                pong()
        elif gameChoice == "back":
            break
        else:
            print("Unknown command.")

def school_menu(username):
    while True:
        render_school(username)

        schoolChoice = input(">> ").lower().strip()

        if schoolChoice == "expenses tracker":
            expenseTracker()
        elif schoolChoice == "homework list":
            homeworkList()
        else:
            print("Unknown Command")



def homepage(username):
    new_window(username)
    while True:
        render_home(username)

        homeChoice = input(">> ").lower().strip()
        if homeChoice == "files":
            files()
            new_window(username)


        # SYSTEM
        elif homeChoice == "system specs":
            systemSpecs()
            new_window(username)
        
        elif homeChoice == "games":
            new_window(username)
            games_menu(username)
            new_window(username)

        elif homeChoice == "school":
            new_window(username)
            school_menu(username)
            new_window(username)

        # OTHER
        elif homeChoice == "calculator":
            calculator()
            new_window(username)

        elif homeChoice == "notes":
            notes()
            new_window(username)

        elif homeChoice == "music":
            music()
            new_window(username)


        # EXIT 
        elif homeChoice == "logout":
            break
        elif homeChoice == "shutdown":
            shutdown()

# BOOT
def boot_screen():
    art = r"""
██████╗ ██╗   ██╗ ██████╗ ███████╗███████╗
██╔══██╗╚██╗ ██╔╝██╔═══██╗██╔════╝██╔════╝
██████╔╝ ╚████╔╝ ██║   ██║███████╗███████╗
██╔═══╝   ╚██╔╝  ██║   ██║╚════██║╚════██║
██║        ██║   ╚██████╔╝███████║███████║
╚═╝        ╚═╝    ╚═════╝ ╚══════╝╚══════╝

    PyOSS - Python Operating System (S)
-------------------------------------
     Welcome to your CLI world <3

    ------- statuzs on dc --------- 
"""
    print(Fore.CYAN + center_block(art))


def new_window(username):
    print("\033c", end="")  # Clear screen


while True:
    new_window("Guest")
    boot_screen()
    print_centered("LOGIN SCREEN")
    print_centered()
    print_centered("New user or existing user?")
    print_centered()
    print_centered("If you are an existing user, please enter your username.")
    print_centered()
    print_centered("Press Q to shutdown")
    print_centered()

    choice = input(center_input_prompt("> ")).strip()

    if choice == "new":
        user = register_user()
        homepage(user)

    elif choice == "q":
        break

    else:
        users = load_users()
        username = choice.strip()

        if username in users:
            password = getpass.getpass(center_input_prompt("Password: "))

            if users[username] == password:
                success_sound()
                homepage(username)
            else:
                error_sound()
                print("Wrong password.")
        else:
            error_sound()
            print("User not found.")